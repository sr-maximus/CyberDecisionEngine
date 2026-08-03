from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
FRAMEWORK_DIR = ROOT / "data" / "frameworks"
SCENARIO_DIR = ROOT / "data" / "scenarios"


def worksheet_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(cell).strip() if cell is not None else "" for cell in next(iterator)]
    rows: list[dict[str, Any]] = []
    for raw in iterator:
        row = {headers[index]: value for index, value in enumerate(raw) if index < len(headers)}
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return rows


def load_disarm() -> dict[str, Any]:
    path = FRAMEWORK_DIR / "DISARM_FRAMEWORKS_MASTER.xlsx"
    tactics = {
        str(row.get("disarm_id")): {
            "id": str(row.get("disarm_id")),
            "name": str(row.get("name") or ""),
            "phase_id": str(row.get("phase_id") or ""),
            "summary": str(row.get("summary") or ""),
        }
        for row in worksheet_rows(path, "tactics")
        if row.get("disarm_id") and row.get("name")
    }
    techniques = []
    for row in worksheet_rows(path, "techniques"):
        disarm_id = str(row.get("disarm_id") or "")
        name = str(row.get("name") or "")
        if not disarm_id.startswith("T") or not name:
            continue
        tactic_id = str(row.get("tactic_id") or "")
        usable = str(row.get("Usable") or "").strip().lower()
        techniques.append(
            {
                "id": disarm_id,
                "name": name,
                "summary": str(row.get("summary") or ""),
                "tactic_id": tactic_id,
                "tactic": tactics.get(tactic_id, {}).get("name", tactic_id),
                "usable": usable == "yes",
            }
        )
    payload = {
        "source": "DISARMFoundation/DISARMframeworks-20-observable",
        "source_url": "https://github.com/DISARMFoundation/DISARMframeworks-20-observable",
        "tactics": list(tactics.values()),
        "techniques": techniques,
    }
    (FRAMEWORK_DIR / "disarm_observable.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def load_attack() -> list[dict[str, Any]]:
    path = FRAMEWORK_DIR / "mitre_attack_enterprise.json"
    bundle = json.loads(path.read_text(encoding="utf-8"))
    rows = []
    for obj in bundle.get("objects", []):
        if obj.get("type") != "attack-pattern" or obj.get("revoked") or obj.get("x_mitre_deprecated"):
            continue
        external_id = ""
        for ref in obj.get("external_references", []):
            if ref.get("source_name") == "mitre-attack":
                external_id = ref.get("external_id", "")
                break
        tactics = [phase.get("phase_name", "") for phase in obj.get("kill_chain_phases", []) if phase.get("phase_name")]
        rows.append(
            {
                "id": external_id or obj.get("id", ""),
                "name": obj.get("name", ""),
                "tactics": tactics,
                "description": obj.get("description", ""),
            }
        )
    return rows


def load_simple_json(path: Path) -> list[dict[str, str]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    return [{"id": key, "name": value} for key, value in data.items()]


def load_f3() -> list[dict[str, Any]]:
    records = json.loads((FRAMEWORK_DIR / "mitre_f3_v1_1.json").read_text(encoding="utf-8"))
    return [item for item in records if isinstance(item, dict) and item.get("tactic") is not True]


def scenario_record(
    index: int,
    attack: dict[str, Any],
    disarm: dict[str, Any],
    d3fend: dict[str, str],
    atlas: dict[str, str],
    sector: str,
) -> dict[str, Any]:
    f3 = attack.get("f3") or {}
    return {
        "id": f"CDE-SCN-{index:04d}",
        "status": "preventive_template",
        "sector": sector,
        "title_es": f"{attack['name']} con narrativa DISARM: {disarm['name']}",
        "title_en": f"{attack['name']} with DISARM narrative: {disarm['name']}",
        "frameworks": {
            "attack": {"id": attack["id"], "name": attack["name"], "tactics": attack.get("tactics", [])[:3]},
            "disarm": {"id": disarm["id"], "name": disarm["name"], "tactic": disarm.get("tactic", "")},
            "d3fend": d3fend,
            "atlas": atlas,
            "f3": f3,
        },
        "scores": {
            "likelihood": 0.0,
            "impact": 0.0,
            "inherent_risk": 0.0,
            "control_effectiveness": 0.0,
            "residual_risk": 0.0,
            "geographic_relevance": 0.0,
        },
        "math": {
            "z": 0.0,
            "formula": "not_calculated_until_current_run_evidence_is_validated",
            "variables": {},
        },
        "recommendation_es": f"Si la evidencia actual coincide de forma explícita, validar el activo y contrastar la opción defensiva {d3fend['id']} {d3fend['name']} antes de decidir tratamiento.",
        "recommendation_en": f"If current evidence matches explicitly, validate the asset and test defensive option {d3fend['id']} {d3fend['name']} before deciding treatment.",
        "strategic_question_es": "¿La narrativa puede alterar confianza, continuidad, fraude, reputación o presión regulatoria en los mercados o países del alcance?",
        "strategic_question_en": "Can the narrative alter trust, continuity, fraud, reputation or regulatory pressure in the markets or countries under scope?",
    }


def f3_scenario_record(
    index: int,
    technique: dict[str, Any],
    d3fend: dict[str, str],
) -> dict[str, Any]:
    attack = (
        {"id": technique["id"], "name": technique["name"], "tactics": technique.get("tactics", [])}
        if technique.get("isAttack")
        else {"id": "", "name": "", "tactics": []}
    )
    return {
        "id": f"CDE-SCN-{index:04d}",
        "status": "preventive_template",
        "sector": "all",
        "title_es": f"Conducta antifraude F3: {technique['name']}",
        "title_en": f"F3 fraud behavior: {technique['name']}",
        "frameworks": {
            "attack": attack,
            "disarm": {"id": "", "name": "", "tactic": ""},
            "d3fend": d3fend,
            "atlas": {"id": "", "name": ""},
            "f3": {
                "id": technique["id"],
                "name": technique["name"],
                "tactics": technique.get("tactics", []),
                "isAttack": bool(technique.get("isAttack")),
            },
        },
        "scores": {
            "likelihood": 0.0,
            "impact": 0.0,
            "inherent_risk": 0.0,
            "control_effectiveness": 0.0,
            "residual_risk": 0.0,
            "geographic_relevance": 0.0,
        },
        "math": {
            "z": 0.0,
            "formula": "not_calculated_until_current_run_evidence_is_validated",
            "variables": {},
        },
        "recommendation_es": (
            f"Si la evidencia de la corrida mapea explícitamente a F3 {technique['id']}, "
            f"contrastar el comportamiento y validar controles defensivos relacionados con "
            f"{d3fend['id']} {d3fend['name']} antes de decidir tratamiento."
        ),
        "recommendation_en": (
            f"If current-run evidence maps explicitly to F3 {technique['id']}, "
            f"contrast the behavior and validate defensive controls related to "
            f"{d3fend['id']} {d3fend['name']} before deciding treatment."
        ),
        "strategic_question_es": (
            "¿La evidencia actual demuestra una conducta antifraude compatible y qué validación "
            "adicional separa una señal preventiva de un fraude confirmado?"
        ),
        "strategic_question_en": (
            "Does current evidence support a compatible fraud behavior and what additional "
            "validation separates a preventive signal from confirmed fraud?"
        ),
    }


def framework_scenario_record(
    index: int,
    framework: str,
    technique: dict[str, Any],
    f3_overlap: dict[str, Any] | None = None,
) -> dict[str, Any]:
    empty_attack = {"id": "", "name": "", "tactics": []}
    empty_disarm = {"id": "", "name": "", "tactic": ""}
    empty_d3fend = {"id": "", "name": ""}
    empty_atlas = {"id": "", "name": ""}
    empty_f3 = {"id": "", "name": "", "tactics": [], "isAttack": False}
    framework_entries: dict[str, dict[str, Any]] = {
        "attack": {
            "attack": {
                "id": technique.get("id", ""),
                "name": technique.get("name", ""),
                "tactics": technique.get("tactics", []),
            },
            "disarm": empty_disarm,
            "d3fend": empty_d3fend,
            "atlas": empty_atlas,
            "f3": f3_overlap or empty_f3,
        },
        "disarm": {
            "attack": empty_attack,
            "disarm": {
                "id": technique.get("id", ""),
                "name": technique.get("name", ""),
                "tactic": technique.get("tactic", ""),
            },
            "d3fend": empty_d3fend,
            "atlas": empty_atlas,
            "f3": empty_f3,
        },
        "atlas": {
            "attack": empty_attack,
            "disarm": empty_disarm,
            "d3fend": empty_d3fend,
            "atlas": {"id": technique.get("id", ""), "name": technique.get("name", "")},
            "f3": empty_f3,
        },
        "f3": {
            "attack": (
                {
                    "id": technique.get("id", ""),
                    "name": technique.get("name", ""),
                    "tactics": technique.get("tactics", []),
                }
                if technique.get("isAttack")
                else empty_attack
            ),
            "disarm": empty_disarm,
            "d3fend": empty_d3fend,
            "atlas": empty_atlas,
            "f3": {
                "id": technique.get("id", ""),
                "name": technique.get("name", ""),
                "tactics": technique.get("tactics", []),
                "isAttack": bool(technique.get("isAttack")),
            },
        },
    }
    framework_labels = {
        "attack": "MITRE ATT&CK",
        "disarm": "DISARM",
        "atlas": "MITRE ATLAS",
        "f3": "MITRE F3",
    }
    framework_label = framework_labels[framework]
    technique_id = str(technique.get("id") or "")
    technique_name = str(technique.get("name") or technique_id)
    return {
        "id": f"CDE-SCN-{index:04d}",
        "status": "preventive_template",
        "sector": "all",
        "title_es": f"{framework_label}: {technique_id} {technique_name}".strip(),
        "title_en": f"{framework_label}: {technique_id} {technique_name}".strip(),
        "frameworks": framework_entries[framework],
        "scores": {
            "likelihood": 0.0,
            "impact": 0.0,
            "inherent_risk": 0.0,
            "control_effectiveness": 0.0,
            "residual_risk": 0.0,
            "geographic_relevance": 0.0,
        },
        "math": {
            "z": 0.0,
            "formula": "not_calculated_until_current_run_evidence_is_validated",
            "variables": {},
        },
        "recommendation_es": (
            f"Si evidencia directa o validada de la corrida activa {framework_label} {technique_id}, "
            "contrastar el comportamiento con el activo afectado y con los controles de referencia "
            "mapeados antes de decidir tratamiento."
        ),
        "recommendation_en": (
            f"If direct or validated current-run evidence activates {framework_label} {technique_id}, "
            "compare the behavior with the affected asset and mapped reference controls before "
            "deciding treatment."
        ),
        "strategic_question_es": (
            "¿La evidencia de la corrida satisface el criterio explícito del marco y mantiene relación "
            "directa con el dominio, activo o entidad del alcance?"
        ),
        "strategic_question_en": (
            "Does current-run evidence satisfy the framework's explicit criterion and retain a direct "
            "relationship to the in-scope domain, asset or entity?"
        ),
    }


def build_scenarios() -> dict[str, Any]:
    disarm = load_disarm()
    attack = load_attack()
    d3fend = load_simple_json(FRAMEWORK_DIR / "mitre_d3fend_minimal.json")
    atlas = load_simple_json(FRAMEWORK_DIR / "mitre_atlas_minimal.json")
    f3 = load_f3()
    usable_disarm = [item for item in disarm["techniques"] if item["usable"]] or disarm["techniques"]
    records: list[dict[str, Any]] = []
    index = 1
    f3_by_attack_id = {
        item["id"]: item
        for item in f3
        if item.get("isAttack") and str(item.get("id", "")).startswith("T")
    }
    for framework, techniques in (
        ("attack", attack),
        ("disarm", usable_disarm),
        ("atlas", atlas),
        ("f3", f3),
    ):
        for technique in techniques:
            overlap = f3_by_attack_id.get(str(technique.get("id") or "")) if framework == "attack" else None
            records.append(framework_scenario_record(index, framework, technique, overlap))
            index += 1
    payload = scenario_payload(
        records,
        catalog_counts={
            "MITRE ATT&CK": len(attack),
            "DISARM": len(usable_disarm),
            "MITRE ATLAS": len(atlas),
            "MITRE F3": len(f3),
            "MITRE D3FEND controls": len(d3fend),
        },
    )
    write_scenarios(payload)
    return payload


def scenario_payload(
    records: list[dict[str, Any]],
    catalog_counts: dict[str, int] | None = None,
) -> dict[str, Any]:
    return {
        "generated_by": "CyberDecisionEngine",
        "version": "2026.07.evidence-gated-framework-derived-v2",
        "scenario_count": len(records),
        "catalog_counts": catalog_counts or {},
        "sources": [
            "MITRE ATT&CK Enterprise STIX",
            "MITRE D3FEND minimal local mapping",
            "MITRE ATLAS minimal local mapping",
            "DISARM Foundation DISARM 2.0 Observations Framework",
            "MITRE Fight Fraud Framework (F3) v1.1",
        ],
        "math_model": {
            "es": "La biblioteca deriva una plantilla por técnica o conducta publicada en ATT&CK, ATLAS, DISARM y F3, sin combinaciones aleatorias ni riesgo precalculado. D3FEND y los marcos de gobierno se cruzan como controles de referencia. La corrida solo activa una plantilla cuando evidencia directa, validada o confirmada satisface el criterio explícito.",
            "en": "The library derives one template per published ATT&CK, ATLAS, DISARM and F3 technique or behavior, without random combinations or precomputed risk. D3FEND and governance frameworks are cross-referenced as control references. A run activates a template only when direct, validated or confirmed evidence satisfies the explicit criterion.",
            "formula": "scenario_support = assured_current_run_evidence_only",
        },
        "scenarios": records,
    }


def write_scenarios(payload: dict[str, Any]) -> None:
    SCENARIO_DIR.mkdir(parents=True, exist_ok=True)
    (SCENARIO_DIR / "cyber_scenario_library.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    payload = build_scenarios()
    print(f"Generated {payload['scenario_count']} scenarios")
